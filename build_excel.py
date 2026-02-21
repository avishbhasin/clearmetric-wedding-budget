"""
ClearMetric Wedding Budget Planner — Premium Excel Template
Product T9 for Gumroad ($14.99)

3 Sheets:
  1. Budget Planner — categories with budgeted, actual, variance
  2. Vendor Tracker — vendor name, category, quoted price, deposit paid, balance, due date, notes
  3. How To Use — instructions

Design: Rose/Blush palette (#8E3553 primary, #6C2742 dark, #F9E4EC input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# ============================================================
# DESIGN SYSTEM — Rose/Blush
# ============================================================
PRIMARY = "8E3553"
DARK = "6C2742"
WHITE = "FFFFFF"
INPUT_TINT = "F9E4EC"
LIGHT_BG = "FDF5F8"
MED_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"
RED = "E74C3C"
LIGHT_RED = "FDEDEC"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="E8D8D8", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color=DARK)
FONT_INPUT = Font(name="Calibri", size=12, color=DARK, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color=DARK)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=DARK)
FONT_SMALL = Font(name="Calibri", size=9, color=MED_GRAY, italic=True)

FILL_PRIMARY = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_TINT, end_color=INPUT_TINT, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
FILL_RED = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY),
    right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY),
    bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_PRIMARY
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_PRIMARY
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# BUDGET CATEGORIES
# ============================================================
BUDGET_CATEGORIES = [
    "Venue & Catering",
    "Photography/Video",
    "Music/DJ",
    "Flowers & Decor",
    "Attire & Beauty",
    "Stationery & Invites",
    "Transportation",
    "Favors & Gifts",
    "Officiant",
    "Miscellaneous/Buffer",
]


# ============================================================
# SHEET 1: BUDGET PLANNER
# ============================================================
def build_budget_planner(ws):
    ws.title = "Budget Planner"
    ws.sheet_properties.tabColor = PRIMARY
    cols(ws, {"A": 2, "B": 28, "C": 14, "D": 14, "E": 14})

    for r in range(1, 55):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="WEDDING BUDGET PLANNER")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:E3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="Enter budgeted amounts in gray cells. Track actual vs budgeted as you go.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # Headers
    header_bar(ws, 5, 2, 2, "Category")
    ws.cell(row=5, column=3, value="Budgeted").font = FONT_HEADER
    ws.cell(row=5, column=3).fill = FILL_PRIMARY
    ws.cell(row=5, column=3).border = THIN
    ws.cell(row=5, column=3).alignment = ALIGN_C
    ws.cell(row=5, column=4, value="Actual").font = FONT_HEADER
    ws.cell(row=5, column=4).fill = FILL_PRIMARY
    ws.cell(row=5, column=4).border = THIN
    ws.cell(row=5, column=4).alignment = ALIGN_C
    ws.cell(row=5, column=5, value="Variance").font = FONT_HEADER
    ws.cell(row=5, column=5).fill = FILL_PRIMARY
    ws.cell(row=5, column=5).border = THIN
    ws.cell(row=5, column=5).alignment = ALIGN_C

    # Default budgeted amounts (industry % of $30k)
    DEFAULTS = {
        "Venue & Catering": 13500,
        "Photography/Video": 3600,
        "Music/DJ": 2100,
        "Flowers & Decor": 2400,
        "Attire & Beauty": 2400,
        "Stationery & Invites": 900,
        "Transportation": 900,
        "Favors & Gifts": 600,
        "Officiant": 300,
        "Miscellaneous/Buffer": 3300,
    }

    for i, cat in enumerate(BUDGET_CATEGORIES):
        r = 6 + i
        ws.cell(row=r, column=2, value=cat).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).alignment = ALIGN_L

        ws.cell(row=r, column=3, value=DEFAULTS.get(cat, 0)).fill = FILL_INPUT
        ws.cell(row=r, column=3).font = FONT_INPUT
        ws.cell(row=r, column=3).number_format = "$#,##0"
        ws.cell(row=r, column=3).border = THIN
        ws.cell(row=r, column=3).alignment = ALIGN_R

        ws.cell(row=r, column=4, value=0).fill = FILL_INPUT
        ws.cell(row=r, column=4).font = FONT_INPUT
        ws.cell(row=r, column=4).number_format = "$#,##0"
        ws.cell(row=r, column=4).border = THIN
        ws.cell(row=r, column=4).alignment = ALIGN_R

        ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
        ws.cell(row=r, column=5).font = FONT_BOLD
        ws.cell(row=r, column=5).number_format = "$#,##0"
        ws.cell(row=r, column=5).border = THIN
        ws.cell(row=r, column=5).alignment = ALIGN_R

    # Total row
    r_total = 6 + len(BUDGET_CATEGORIES)
    ws.cell(row=r_total, column=2, value="TOTAL").font = FONT_BOLD
    ws.cell(row=r_total, column=2).fill = FILL_GRAY
    ws.cell(row=r_total, column=2).border = THIN
    ws.cell(row=r_total, column=3, value=f"=SUM(C6:C{r_total-1})").font = FONT_BOLD
    ws.cell(row=r_total, column=3).number_format = "$#,##0"
    ws.cell(row=r_total, column=3).border = THIN
    ws.cell(row=r_total, column=4, value=f"=SUM(D6:D{r_total-1})").font = FONT_BOLD
    ws.cell(row=r_total, column=4).number_format = "$#,##0"
    ws.cell(row=r_total, column=4).border = THIN
    ws.cell(row=r_total, column=5, value=f"=D{r_total}-C{r_total}").font = FONT_BOLD
    ws.cell(row=r_total, column=5).number_format = "$#,##0"
    ws.cell(row=r_total, column=5).border = THIN

    ws.protection.sheet = True
    for r in range(6, r_total):
        for c in [3, 4]:
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: VENDOR TRACKER
# ============================================================
def build_vendor_tracker(wb):
    ws = wb.create_sheet("Vendor Tracker")
    ws.sheet_properties.tabColor = "B85C7A"
    cols(ws, {"A": 2, "B": 24, "C": 18, "D": 14, "E": 14, "F": 14, "G": 14, "H": 4})

    for r in range(1, 45):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 9):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:H1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:H2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="VENDOR TRACKER").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:H3")
    ws.cell(row=3, column=2, value="Track vendor quotes, deposits, balance, and due dates.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    # Headers
    headers = ["Vendor Name", "Category", "Quoted Price", "Deposit Paid", "Balance", "Due Date", "Notes"]
    for i, h in enumerate(headers):
        c = 2 + i
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_C
        cell.border = THIN

    # Sample rows (5 vendors)
    sample_vendors = [
        ("Venue Name", "Venue & Catering", 15000, 3000, 12000, "2025-06-15", "50% deposit due"),
        ("Photographer", "Photography/Video", 3500, 500, 3000, "2025-06-01", "8-hour package"),
        ("Florist", "Flowers & Decor", 2200, 0, 2200, "2025-05-20", ""),
        ("DJ", "Music/DJ", 1800, 400, 1400, "2025-06-10", ""),
        ("Caterer", "Venue & Catering", 8500, 0, 8500, "2025-05-01", "Per guest"),
    ]

    for i, row_data in enumerate(sample_vendors):
        r = 6 + i
        # Vendor, Category, Quoted, Deposit, Due Date, Notes (skip Balance - it's a formula)
        cols_map = [2, 3, 4, 5, 7, 8]  # B, C, D, E, G, H (skip F=Balance)
        vals = [row_data[0], row_data[1], row_data[2], row_data[3], row_data[5], row_data[6]]
        for j, (col_idx, val) in enumerate(zip(cols_map, vals)):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = THIN
            cell.fill = FILL_INPUT
            cell.font = FONT_INPUT
            if col_idx in [4, 5]:  # quoted, deposit
                cell.number_format = "$#,##0"
                cell.alignment = ALIGN_R
            elif col_idx == 7:  # due date
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = ALIGN_C
            else:
                cell.alignment = ALIGN_L

        # Balance formula: = Quoted - Deposit (D=Quoted, E=Deposit)
        ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        ws.cell(row=r, column=6).font = FONT_BOLD
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=6).fill = FILL_WHITE
        ws.cell(row=r, column=6).border = THIN
        ws.cell(row=r, column=6).alignment = ALIGN_R

    # Add empty rows for more vendors
    for i in range(5, 15):
        r = 6 + i
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = THIN
            if c != 6:
                cell.fill = FILL_INPUT
        ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=6).border = THIN

    ws.protection.sheet = True
    for r in range(6, 21):
        for c in [2, 3, 4, 5, 7, 8]:  # vendor, category, quoted, deposit, due date, notes
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = MED_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE WEDDING BUDGET PLANNER")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Budget Planner' tab and enter your budgeted amounts in the GRAY cells",
            "2. As you pay vendors, update the 'Actual' column. Variance updates automatically.",
            "3. Use the 'Vendor Tracker' tab to track quotes, deposits, balance, and due dates.",
        ]),
        ("BUDGET PLANNER", [
            "Budgeted: Your planned spend for each category (industry avg: 45% venue/catering, 12% photo, etc.)",
            "Actual: What you've actually spent. Update as you go.",
            "Variance: Actual minus Budgeted. Negative = under budget, Positive = over budget.",
        ]),
        ("VENDOR TRACKER", [
            "Vendor Name: Name of the vendor or service provider",
            "Category: Venue & Catering, Photography/Video, etc.",
            "Quoted Price: Total agreed price",
            "Deposit Paid: Amount already paid",
            "Balance: Quoted - Deposit (auto-calculated)",
            "Due Date: When payment is due",
            "Notes: Any reminders or special terms",
        ]),
        ("TIPS", [
            "Venue & Catering typically eats 40-50% of the budget. Negotiate early.",
            "Photography and flowers are often over budget. Get 3 quotes.",
            "Set aside 10-15% for misc/buffer — unexpected costs always come up.",
            "Track deposits carefully — you'll have many partial payments due at different times.",
        ]),
        ("IMPORTANT NOTES", [
            "This planner is for educational purposes only. Not financial advice.",
            "Rates and availability vary by region and season.",
            "© 2026 ClearMetric. For personal use only.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=DARK)
        ws.cell(row=r, column=2).fill = PatternFill(start_color=INPUT_TINT, end_color=INPUT_TINT, fill_type="solid")
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color=DARK)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Budget Planner sheet...")
    build_budget_planner(ws)

    print("Building Vendor Tracker sheet...")
    build_vendor_tracker(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "output",
        "ClearMetric-Wedding-Budget-Planner.xlsx",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
